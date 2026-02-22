"""
CSV Reporter - Professional extraction reports for quality control
V3.1 - Generate comprehensive CSV reports for audit and verification
"""

import csv
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel reports disabled")


@dataclass
class ExtractionRecord:
    """Single extraction record for CSV export"""
    timestamp: str
    pdf_filename: str
    page_number: int
    header_extracted: str
    confidence_score: int
    ocr_method: str
    processing_time_ms: float
    render_scale: float
    status: str  # success, low_confidence, error
    error_message: str = ""
    quality_flags: str = ""
    split_group: str = ""  # Which split PDF it belongs to
    output_filename: str = ""
    
    def to_dict(self) -> dict:
        """Convert to dictionary for CSV writing"""
        return asdict(self)


class CSVReporter:
    """
    Generate professional Excel reports for extraction quality control
    
    Features:
    - Per-page extraction details with color highlighting
    - Summary statistics at top
    - Confidence scores
    - Error tracking
    - Split group mapping
    - Daily/per-file reports
    """
    
    def __init__(
        self,
        output_folder: str = 'reports',
        organize_by_date: bool = True,
        append_mode: bool = True,
        use_excel: bool = True
    ):
        """
        Initialize CSV Reporter
        
        Args:
            output_folder: Base folder for reports
            organize_by_date: Create date-based subfolders
            append_mode: Append to existing files (vs overwrite)
        """
        self.output_folder = Path(output_folder)
        self.organize_by_date = organize_by_date
        self.append_mode = append_mode
        self.use_excel = use_excel and EXCEL_AVAILABLE
        
        # Create output folder
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # Track records for batch writing
        self.pending_records: List[ExtractionRecord] = []
        
        # Track first extraction time for filename
        self.first_extraction_time: Optional[datetime] = None
        
        logger.info(f"Reporter initialized: {self.output_folder} (Excel: {self.use_excel})")
    
    def add_extraction(
        self,
        pdf_filename: str,
        page_number: int,
        header_extracted: str,
        confidence_score: int,
        ocr_method: str,
        processing_time_ms: float,
        render_scale: float = 2.0,
        status: str = "success",
        error_message: str = "",
        quality_flags: str = "",
        split_group: str = "",
        output_filename: str = ""
    ):
        """
        Add an extraction record
        
        Args:
            pdf_filename: Source PDF filename
            page_number: Page number (1-based)
            header_extracted: Extracted header text
            confidence_score: Validation score (0-300)
            ocr_method: OCR method used
            processing_time_ms: Processing time in milliseconds
            render_scale: Render scale used
            status: Status (success, low_confidence, error)
            error_message: Error message if failed
            split_group: Which split group (for PDF splitting)
            output_filename: Output filename if split
        """
        # Track first extraction time for filename
        if self.first_extraction_time is None:
            self.first_extraction_time = datetime.now()
        
        record = ExtractionRecord(
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            pdf_filename=pdf_filename,
            page_number=page_number,
            header_extracted=header_extracted,
            confidence_score=confidence_score,
            ocr_method=ocr_method,
            processing_time_ms=round(processing_time_ms, 2),
            render_scale=render_scale,
            status=status,
            error_message=error_message,
            quality_flags=quality_flags,
            split_group=split_group,
            output_filename=output_filename
        )
        
        self.pending_records.append(record)
    
    def flush_to_csv(
        self,
        job_id: Optional[str] = None,
        summary_stats: Optional[Dict] = None
    ) -> Path:
        """
        Write pending records to Excel/CSV file
        
        Args:
            job_id: Optional job ID for filename
            summary_stats: Optional summary statistics to write at top
        
        Returns:
            Path to created file
        """
        if not self.pending_records:
            logger.warning("No records to write")
            return None
        
        # Determine output path
        if self.organize_by_date:
            date_folder = self.output_folder / datetime.now().strftime("%Y-%m-%d")
            date_folder.mkdir(parents=True, exist_ok=True)
            output_path = date_folder
        else:
            output_path = self.output_folder
        
        # Generate filename using first extraction time
        if self.first_extraction_time:
            timestamp = self.first_extraction_time.strftime("%Y%m%d_%H%M%S")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Use Excel if available, otherwise CSV
        if self.use_excel:
            return self._write_excel(output_path, timestamp, summary_stats)
        else:
            return self._write_csv(output_path, timestamp, summary_stats)
    
    def _write_excel(
        self,
        output_path: Path,
        timestamp: str,
        summary_stats: Optional[Dict] = None
    ) -> Path:
        """
        Write Excel report with formatting and highlighting
        
        Args:
            output_path: Output directory
            timestamp: Timestamp for filename
            summary_stats: Summary statistics to write at top
        
        Returns:
            Path to created Excel file
        """
        try:
            excel_filename = f"extraction_report_{timestamp}.xlsx"
            excel_path = output_path / excel_filename
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Extraction Report"
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            error_font = Font(color="FFFFFF", bold=True)
            warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            warning_font = Font(color="000000", bold=True)
            summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            summary_font = Font(bold=True, size=12)
            center_align = Alignment(horizontal="center", vertical="center")
            
            current_row = 1
            
            # Write summary statistics first
            if summary_stats:
                ws.merge_cells(f'A{current_row}:D{current_row}')
                summary_title = ws.cell(current_row, 1, "SUMMARY STATISTICS")
                summary_title.font = Font(bold=True, size=14, color="1F4E78")
                summary_title.alignment = center_align
                current_row += 1
                
                # Write each stat
                for key, value in summary_stats.items():
                    label_cell = ws.cell(current_row, 1, key.replace('_', ' ').title())
                    label_cell.font = summary_font
                    label_cell.fill = summary_fill
                    
                    value_cell = ws.cell(current_row, 2, str(value))
                    value_cell.fill = summary_fill
                    
                    current_row += 1
                
                current_row += 1  # Empty row separator
            
            # Write header row for data
            fieldnames = list(self.pending_records[0].to_dict().keys())
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                cell = ws.cell(current_row, col_idx, fieldname.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            current_row += 1
            data_start_row = current_row
            
            # Write data rows with conditional formatting
            for record in self.pending_records:
                row_data = record.to_dict()
                
                for col_idx, fieldname in enumerate(fieldnames, start=1):
                    cell = ws.cell(current_row, col_idx, row_data[fieldname])
                    
                    # Apply conditional formatting based on status
                    if record.status == 'error':
                        cell.fill = error_fill
                        cell.font = error_font
                    elif record.status == 'low_confidence':
                        cell.fill = warning_fill
                        cell.font = warning_font
                
                current_row += 1
            
            # Auto-size columns
            for col_idx in range(1, len(fieldnames) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(excel_path)
            
            logger.info(f"Excel report written: {excel_path} ({len(self.pending_records)} records)")
            
            # Clear pending records and reset time
            self.pending_records.clear()
            self.first_extraction_time = None
            
            return excel_path
        
        except Exception as e:
            logger.error(f"Failed to write Excel report: {e}", exc_info=True)
            # Fallback to CSV
            return self._write_csv(output_path, timestamp, summary_stats)
    
    def _write_csv(
        self,
        output_path: Path,
        timestamp: str,
        summary_stats: Optional[Dict] = None
    ) -> Path:
        """
        Write CSV report (fallback when Excel not available)
        
        Args:
            output_path: Output directory
            timestamp: Timestamp for filename
            summary_stats: Summary statistics to write at top
        
        Returns:
            Path to created CSV file
        """
        try:
            csv_filename = f"extraction_report_{timestamp}.csv"
            csv_path = output_path / csv_filename
            
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                # Write summary first
                if summary_stats:
                    f.write("# SUMMARY STATISTICS\n")
                    for key, value in summary_stats.items():
                        f.write(f"# {key},{value}\n")
                    f.write("\n")
                
                # Write data
                if self.pending_records:
                    fieldnames = list(self.pending_records[0].to_dict().keys())
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for record in self.pending_records:
                        writer.writerow(record.to_dict())
            
            logger.info(f"CSV report written: {csv_path} ({len(self.pending_records)} records)")
            
            # Clear pending records and reset time
            self.pending_records.clear()
            self.first_extraction_time = None
            
            return csv_path
        
        except Exception as e:
            logger.error(f"Failed to write CSV report: {e}", exc_info=True)
            return None
    
    def generate_summary_report(self, csv_path: Path) -> Dict:
        """
        Generate summary statistics from CSV
        
        Args:
            csv_path: Path to CSV file
        
        Returns:
            Dictionary of summary statistics
        """
        try:
            records = []
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if not row.get('timestamp'):  # Skip summary rows
                        break
                    records.append(row)
            
            if not records:
                return {}
            
            # Calculate statistics
            total_pages = len(records)
            success_count = sum(1 for r in records if r['status'] == 'success')
            error_count = sum(1 for r in records if r['status'] == 'error')
            low_confidence_count = sum(1 for r in records if r['status'] == 'low_confidence')
            
            # Confidence scores
            scores = [int(r['confidence_score']) for r in records if r['confidence_score'].isdigit()]
            avg_score = sum(scores) / len(scores) if scores else 0
            
            # Processing times
            times = [float(r['processing_time_ms']) for r in records if r['processing_time_ms']]
            avg_time = sum(times) / len(times) if times else 0
            
            # Unique headers
            unique_headers = len(set(r['header_extracted'] for r in records if r['header_extracted']))
            
            summary = {
                'total_pages': total_pages,
                'success_count': success_count,
                'error_count': error_count,
                'low_confidence_count': low_confidence_count,
                'success_rate': f"{(success_count / total_pages * 100):.1f}%",
                'avg_confidence_score': round(avg_score, 1),
                'avg_processing_time_ms': round(avg_time, 2),
                'unique_headers': unique_headers,
                'unique_split_groups': len(set(r['split_group'] for r in records if r['split_group']))
            }
            
            return summary
        
        except Exception as e:
            logger.error(f"Failed to generate summary: {e}")
            return {}
    
    def create_error_report(self, report_path: Path, error_records_data: List[ExtractionRecord]) -> Optional[Path]:
        """
        Create a separate report with only errors and low-confidence extractions
        
        Args:
            report_path: Main report path (used for determining error report path)
            error_records_data: List of error/low-confidence records
        
        Returns:
            Path to error report file
        """
        try:
            if not error_records_data:
                logger.info("No errors or low-confidence extractions found")
                return None
            
            # Determine error report path
            if report_path.suffix == '.xlsx':
                error_path = report_path.parent / f"errors_{report_path.stem}.xlsx"
                return self._write_error_excel(error_path, error_records_data)
            else:
                error_path = report_path.parent / f"errors_{report_path.name}"
                return self._write_error_csv(error_path, error_records_data)
        
        except Exception as e:
            logger.error(f"Failed to create error report: {e}", exc_info=True)
            return None
    
    def _write_error_excel(self, error_path: Path, error_records: List[ExtractionRecord]) -> Path:
        """Write error report in Excel format"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error Report"
            
            # Define styles
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            error_font = Font(color="FFFFFF", bold=True)
            warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            warning_font = Font(color="000000", bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Write header
            fieldnames = list(error_records[0].to_dict().keys())
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                cell = ws.cell(1, col_idx, fieldname.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Write error records
            for row_idx, record in enumerate(error_records, start=2):
                row_data = record.to_dict()
                
                for col_idx, fieldname in enumerate(fieldnames, start=1):
                    cell = ws.cell(row_idx, col_idx, row_data[fieldname])
                    
                    # Apply highlighting
                    if record.status == 'error':
                        cell.fill = error_fill
                        cell.font = error_font
                    elif record.status == 'low_confidence':
                        cell.fill = warning_fill
                        cell.font = warning_font
            
            # Auto-size columns
            for col_idx in range(1, len(fieldnames) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(error_path)
            logger.info(f"Error report created: {error_path} ({len(error_records)} issues)")
            return error_path
        
        except Exception as e:
            logger.error(f"Failed to write error Excel: {e}", exc_info=True)
            return None
    
    def _write_error_csv(self, error_path: Path, error_records: List[ExtractionRecord]) -> Path:
        """Write error report in CSV format"""
        try:
            with open(error_path, 'w', newline='', encoding='utf-8-sig') as f:
                fieldnames = list(error_records[0].to_dict().keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for record in error_records:
                    writer.writerow(record.to_dict())
            
            logger.info(f"Error report created: {error_path} ({len(error_records)} issues)")
            return error_path
        
        except Exception as e:
            logger.error(f"Failed to write error CSV: {e}", exc_info=True)
            return None
    
    def create_daily_summary(self) -> Path:
        """
        Create a daily summary CSV aggregating all jobs
        
        Returns:
            Path to daily summary CSV
        """
        today = datetime.now().strftime("%Y-%m-%d")
        date_folder = self.output_folder / today
        
        if not date_folder.exists():
            logger.warning(f"No reports found for {today}")
            return None
        
        # Collect all CSV files for today
        csv_files = list(date_folder.glob("extraction_report_*.csv"))
        
        if not csv_files:
            logger.warning(f"No extraction reports found for {today}")
            return None
        
        # Aggregate data
        all_records = []
        for csv_file in csv_files:
            try:
                with open(csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if row.get('timestamp'):
                            all_records.append(row)
            except Exception as e:
                logger.error(f"Failed to read {csv_file}: {e}")
        
        if not all_records:
            return None
        
        # Write daily summary
        summary_path = date_folder / f"daily_summary_{today}.csv"
        
        with open(summary_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=all_records[0].keys())
            writer.writeheader()
            writer.writerows(all_records)
        
        logger.info(f"Daily summary created: {summary_path} ({len(all_records)} records)")
        
        # Generate summary stats
        summary_stats = self.generate_summary_report(summary_path)
        
        # Append summary
        with open(summary_path, 'a', encoding='utf-8-sig') as f:
            f.write("\n\n# Daily Summary Statistics\n")
            for key, value in summary_stats.items():
                f.write(f"# {key},{value}\n")
        
        return summary_path
